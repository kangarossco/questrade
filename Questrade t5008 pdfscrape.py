#############################################################################################################
# Script Name       : Questrade t5008 pdfscrape.py                                                          #
# script version    : 1                                                                                     #
# Author            : Ross McKinnon                                                                         #
# contact:          : srossmckinnon@gmail.com                                                               #
# Date              : February 28, 2021                                                                     #
# Python Version    : python 3.9.0 windows                                                                  #
# Description       : parses trading summary from questrade t5008 pdf, converts information to an .xlsx     #
#############################################################################################################

#Import modules
import PyPDF2
import openpyxl
import os

#for year identifcation at end of page
import datetime

#location of pdf source data and location/name of xlworkbook that will be created
pdffilepath = r"C:\Users\kangarossco\Downloads\Questrade T5008 Original.pdf"
xldestination = r"C:\Users\kangarossco\Downloads\Questrade T5008 Original.xlsx"

#If troubleshooting, I found it easier to delete the file and start from scratch rather than delete
#the data in the file. Uncomment if you plan on running multiple times
if(os.path.exists(xldestination)):
    os.remove(xldestination)

#Create workbook (may get errors if book exists)
wb = openpyxl.Workbook()
wb.save(filename = xldestination)

#create headers and column widths for consitancy on every page
headers = ['Currency', 'Date', 'Type Code', 'Quantity', 'Identification', 'ISIN/CUSIP', 'Cost/Book Value', 'Proceeds of Disposition']
col_widths = [8.5, 10, 10.5, 8.5, 60, 11, 14, 20]

#import entire pdf object and read it
pdfFileObj = open(pdffilepath, 'rb')
pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

#meat and potatoes of the script here
#loop through each pdf page one at a time
for page in range(0,pdfReader.numPages-1):

    #create sheet in excel orkbook and name the tab after the page we are on
    sheet = wb.create_sheet("Page " + str(page+1))

    #add headers and adjust column sizes for this particular tab
    #the commented out line is for auto-adjusting the columns, not reccomended for format before the data is in
    for tab in range(len(headers)):
        sheet.cell(row = 1, column = tab + 1, value = headers[tab])
        #sheet.column_dimensions[get_column_letter(tab+1)].auto_size = True
        sheet.column_dimensions[openpyxl.utils.get_column_letter(tab+1)].width = col_widths[tab]

    #import pdf object and extract text as string
    #convet string to a list delimited by '\n' to seperate elements
    pageObj = pdfReader.getPage(page)
    mytext = pageObj.extractText()
    z = mytext.split('\n')


    #Find the beginning of the useful information by looking for either 'CAD' or 'USD' at the beginning of the string
    #this will compare the currency indices and use the one that's closer to the beginning of the page
    #of course this wouldn't work if you are trading with more than USD and CAD currencies. In this case the start of the
    #useful information could be found by (RealStart = z.index("Produits de") + n) or similar

    #if the currency isn't found it will return a ValueError, so this must be accomodated for considering there is no
    #guarantee both currencies will have trades on each page
    try:    
        USDStart = z.index("USD")
    except ValueError:
        USDStart = None

    try:
        CADStart = z.index("CAD")
    except ValueError:
        CADStart = None

    #compare currencies and see which one is closer to the top
    #that will be our starting point
    if USDStart != None and CADStart != None:
        if USDStart < CADStart:
            RealStart = USDStart
        elif CADStart < USDStart:
            RealStart = CADStart
    elif USDStart == None:
        RealStart = CADStart
    elif CADStart == None:
        RealStart = USDStart

    #find the end of the useful information with a common delimiter, each page ends with code "RC-19-1446"
    #update list with common ending. The 19 in the middle is the last two digits of last year's date.
    #if you're going back several years make sure to update accordingly
    t5008_year = str(datetime.datetime.now().year - 1 - 2000)
    eop_id = "RC-" + t5008_year + "-1446"
    end_of_page = z.index(eop_id)

    #find row numbers by counting all the '$' symbols on each page and dividing by 2
    #there are only two datum in each row with '$' so it works out pretty well.
    # Note: only works cause there are no other '$'s in the preamble. may need extra conditions to work in the future
    row_num = int(mytext.count('$')/2)

    #create new list with only the information needed and set the interation counter to 0
    z = z[RealStart:end_of_page]
    RealStart = 0

    #printIndex is just for printing the data to the excel sheet, and not for cleaning the data
    printIndex = RealStart

    #many nested loops to tackle all the conditions and clean the data properly
    for n in range(int(row_num)):

        #each row should have 2 '$'s, so we keep track with dollarSigns counter
        #proceeds_of_disposition signals the end of each row
        dollarSigns = 0
        proceeds_of_disposition = 0

        #go through 8 elements row by row in the list
        for m in range(len(z[RealStart:RealStart+8])):

            #if you are at the beginning of the row and there is a $ sign 8 elements away it means
            #you are doubled up on the Identificaiton row
            #so we add the second element to the first and delete the second element
            try:
                if m <= 2:
                    if z[RealStart + m + 8].find('$') == 0:
                        z[RealStart + m + 4] = z[RealStart + m + 4] + z[RealStart + m + 4 + 1]
                        z.pop(RealStart + m + 4 + 1)

            #if nothing happens, great
            #but it might throw an error
            except: 
                pass

            #look for options trades by searching for 'PUT' or 'CALL' in the text, these rows will not have ISIN data
            #so an element needs to be added before the first element with $ in it
            #ie, if there is a $ after the cell with CALL/PUT it's only one element of data
            #otherwise it's two
            #this obviously wouldn't work if there was a ticker with letters "P-U-T" or "C-A-L-L"
            #but I'm sure this isn't the case
            if z[RealStart + m].find('PUT') == 0 or z[RealStart + m].find('CALL') == 0:
                if z[RealStart + m + 1].find('$') == 0:
                    z.insert(RealStart + m + 1," ")
                else:
                    z[RealStart + m] = z[RealStart + m] + z[RealStart + m + 1]
                    z[RealStart + m + 1] = " "

            #keep track of how many $ we've been through
            #the .find() function returns -1 if it fails
            if z[RealStart + m].find('$') != -1:
                proceeds_of_disposition = m
                dollarSigns += 1

            #seeing the second $ means the row is finished
            if dollarSigns == 2:
                break
            
        #the new row starts at RealStart by adding the number of elements in the previous row + 1 to adjust
        RealStart = RealStart + proceeds_of_disposition + 1

        #this for loop actually prints the information to the excel sheet using a double for loop
        #we know exactly how many rows there are, and the data should be tidy'd at this point
        for i in range(int(len(z)/8)):
            for j in range(8):
                sheet.cell(row = i + 2, column = j + 1, value = z[printIndex + j + 8 * i])

#get rid of that extra placeholder sheet 
del wb['Sheet']

#save workbook
wb.save(xldestination)

#and you're done!
print('Fin')
